#!/usr/bin/env python3
"""
csv2excel.py - CSV to Excel Converter with Custom Tab Names

This script converts one or more CSV files into a single Excel file, with the option
to specify custom tab names for each CSV. It handles duplicate tab names, applies
filters to all headers, and provides a progress bar for large file processing.

Features:
- Convert multiple CSV files to a single Excel file
- Specify custom tab names for each CSV file
- Automatically handle duplicate tab names
- Apply filters to all headers in the Excel file
- Display a progress bar during processing
- Skip empty or unreadable CSV files
- Provide informative warnings and error messages

Usage:
python csv2excel.py input1.csv "Tab Name 1" input2.csv input3.csv "Tab Name 3" -o output.xlsx

Requirements:
- Python 3.6+
- pandas
- openpyxl
- tqdm
"""

import os
import argparse
from typing import List, Tuple, Set
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

def get_unique_sheet_name(existing_names: Set[str], base_name: str) -> str:
    """
    Generate a unique sheet name by appending a number if the base name already exists.

    Args:
        existing_names (Set[str]): Set of existing sheet names.
        base_name (str): The desired base name for the sheet.

    Returns:
        str: A unique sheet name.
    """
    if base_name not in existing_names:
        return base_name
    i = 1
    while f"{base_name}_{i}" in existing_names:
        i += 1
    return f"{base_name}_{i}"

def csv_to_excel(input_data: List[Tuple[str, str]], output_file: str) -> None:
    """
    Convert CSV files to a single Excel file with custom tab names.

    Args:
        input_data (List[Tuple[str, str]]): List of tuples containing CSV file paths and their corresponding tab names.
        output_file (str): Path to the output Excel file.

    Raises:
        ValueError: If no input data is provided.
    """
    if not input_data:
        raise ValueError("No input data provided")

    total_size = sum(os.path.getsize(csv_path) for csv_path, _ in input_data if os.path.exists(csv_path))
    progress_bar = tqdm(total=total_size, unit='B', unit_scale=True, desc="Processing")

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        existing_names: Set[str] = set()
        for csv_path, tab_name in input_data:
            if not os.path.exists(csv_path):
                print(f"Warning: File {csv_path} does not exist. Skipping.")
                continue
            
            if os.path.isfile(csv_path) and csv_path.endswith('.csv'):
                try:
                    df = pd.read_csv(csv_path)
                except pd.errors.EmptyDataError:
                    print(f"Warning: {csv_path} is empty. Skipping.")
                    continue
                except Exception as e:
                    print(f"Error reading {csv_path}: {str(e)}. Skipping.")
                    continue
            elif os.path.isdir(csv_path):
                print(f"Warning: {csv_path} is a directory. Skipping.")
                continue
            else:
                print(f"Warning: {csv_path} is not a CSV file. Skipping.")
                continue
            
            if not tab_name:
                tab_name = os.path.splitext(os.path.basename(csv_path))[0]
            
            sheet_name = get_unique_sheet_name(existing_names, tab_name)
            existing_names.add(sheet_name)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            progress_bar.update(os.path.getsize(csv_path))

    progress_bar.close()

    # Apply filters to all sheets
    workbook = load_workbook(output_file)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        worksheet.auto_filter.ref = worksheet.dimensions
    
    workbook.save(output_file)

def parse_input(input_args: List[str]) -> List[Tuple[str, str]]:
    """
    Parse input arguments to extract CSV file paths and their corresponding tab names.

    Args:
        input_args (List[str]): List of input arguments from the command line.

    Returns:
        List[Tuple[str, str]]: List of tuples containing CSV file paths and their corresponding tab names.
    """
    input_data = []
    i = 0
    while i < len(input_args):
        csv_path = input_args[i]
        if i + 1 < len(input_args) and not input_args[i + 1].endswith('.csv') and not os.path.isdir(input_args[i + 1]):
            tab_name = input_args[i + 1]
            i += 2
        else:
            tab_name = None
            i += 1
        input_data.append((csv_path, tab_name))
    return input_data

def main() -> None:
    """
    Main function to parse command-line arguments and execute the CSV to Excel conversion.
    """
    parser = argparse.ArgumentParser(description='Convert CSV files to Excel with filtered headers and custom tab names.')
    parser.add_argument('input', nargs='+', help='CSV files or folders, optionally followed by desired tab names')
    parser.add_argument('-o', '--output', required=True, help='Name of the output Excel file')
    
    args = parser.parse_args()
    
    input_data = parse_input(args.input)
    try:
        csv_to_excel(input_data, args.output)
        print(f"Excel file '{args.output}' created successfully.")
        print(f"Processed {len(input_data)} input items.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()